import os
import shutil
import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from threading import Thread
import sys # Import sys module

Update_file_path = None
# Import pywin32 for COM automation
try:
    import win32com.client as win32
    VBA_AUTOMATION_AVAILABLE = True
except ImportError:
    VBA_AUTOMATION_AVAILABLE = False
    print("Warning: 'pywin32' not found. VBA automation will not be available.")
    print("Install with: pip install pywin32")


column_mapping = {
    "Load Number": "LOAD",
    "RO description": "Check flujo",
    "Número de carga externo": "AUTEX N°",
    "Service Level": "Tipo de Servicio",
    "Peso de la carga [kg]": "PESO",
    "Volumen de la carga [m³]": "M3",
    "Medición de la carga [LM]": "SATURACION",
    "Consignor Customer-ID": "COFOR",
    "Consignor Company": "PROVEEDOR",
    "Consignor City": "CIUDAD",
    "Transport Mode": "FLUJO",
    "Means of transport": "TIPO DE VEHICULO COLECTA",
    "Plate Truck": "PLACA COLECTA",
    "Código: Puerto de origen": "CRT",
    "N° de pedido": "NF",
    "Salida del lugar de recogida": "SAÍDA REAL PROVEDOR",
    "Invoice": "FACTURA",
    "TO Comment": "OBSERVACIONES",
    "Latest release date": "LIBERACION EN FRONTERA",
    "Pickup date": "FECHA DE COLECTA iTMS",
    "Delivery date": "FECHA DESCARGA",
    "Llegada al lugar de entrega": "ARRIBO STELLANTIS",
    "ETA": "PREVISION ARRIBO STELLANTIS",
    "ETA al lugar de entrega": "VENTANA ARRIBO STELLANTIS",
    "Service provider Company": "TRANSPORTE"
}

date_columns = {
    "FECHA DE COLECTA iTMS",
    "FECHA DESCARGA",
    "LIBERACION EN FRONTERA",
    "PREVISION ARRIBO STELLANTIS",
    "VENTANA ARRIBO STELLANTIS",
    "ARRIBO STELLANTIS"
}


class ExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Merger & Data Processor")
        self.root.geometry("500x300")
        self.root.resizable(False, False)

        # Style configuration
        self.style = ttk.Style()
        self.style.theme_use('clam')  # 'clam', 'alt', 'default', 'classic'

        # Define colors
        self.primary_color = "#4CAF50"  # Green
        self.secondary_color = "#388E3C" # Darker Green
        self.bg_color = "#ECEFF1"      # Light Grey
        self.text_color = "#263238"     # Dark Blue Grey
        self.accent_color = "#FFC107"   # Amber

        self.style.configure('TFrame', background=self.bg_color)
        self.style.configure('TLabel', background=self.bg_color, foreground=self.text_color, font=('Segoe UI', 10))
        self.style.configure('TButton', background=self.primary_color, foreground='white', font=('Segoe UI', 10, 'bold'), borderwidth=0)
        self.style.map('TButton',
                       background=[('active', self.secondary_color), ('pressed', self.secondary_color)],
                       foreground=[('active', 'white')])
        self.style.configure('TProgressbar', thickness=15, troughcolor=self.bg_color, background=self.primary_color)

        self.root.configure(background=self.bg_color)

        # Main frame
        self.main_frame = ttk.Frame(self.root, padding="30 30 30 30")
        self.main_frame.pack(expand=True, fill="both")

        # Title
        title_label = ttk.Label(self.main_frame, text="Excel Data Merger", font=('Segoe UI', 16, 'bold'), foreground=self.primary_color)
        title_label.pack(pady=(0, 20))

        # Folder selection
        self.label_folder = ttk.Label(self.main_frame, text="Select folder containing 'loading' and 'seguimiento' Excel files:")
        self.label_folder.pack(pady=(10, 5))

        self.btn_browse = ttk.Button(self.main_frame, text="Browse Folder", command=self.browse_folder)
        self.btn_browse.pack(pady=10)

        # Progress bar
        self.progress = ttk.Progressbar(self.main_frame, orient="horizontal", length=350, mode="determinate")
        self.progress.pack(pady=20)

        # Status bar
        self.status_label = ttk.Label(self.root, text="Ready", relief=tk.SUNKEN, anchor=tk.W, font=('Segoe UI', 9), foreground=self.text_color)
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X)

    def update_status(self, message):
        self.status_label.config(text=message)
        self.root.update_idletasks()

    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.update_status(f"Selected folder: {folder_selected}")
            self.progress["value"] = 0
            self.btn_browse.config(state=tk.DISABLED) # Disable button during processing
            Thread(target=self.process_files, args=(folder_selected,)).start()

    def parse_date(self, value):
        if pd.isna(value) or value == "":
            return None

        if isinstance(value, datetime):
            return value

        try:
            return datetime.strptime(str(value).strip(), "%d.%m.%Y")
        except ValueError:
            try:
                return datetime.strptime(str(value).strip(), "%d/%m/%Y")
            except ValueError:
                return value

    def process_files(self, folder_path):
        try:
            self.update_status("Processing files...")
            self.progress["value"] = 10
            self.root.update_idletasks()

            loading_file = next((f for f in os.listdir(folder_path) if "loading" in f.lower()), None)
            seguimento_file = next((f for f in os.listdir(folder_path) if "seguimiento" in f.lower() and not f.startswith("~$")), None)
            Update_file_path = folder_path

            if not loading_file or not seguimento_file:
                messagebox.showerror("Error", "Required files not found ('loading' and 'seguimiento'). Please ensure both are in the selected folder.")
                self.update_status("Error: Required files not found.")
                return

            loading_path = os.path.join(folder_path, loading_file)
            seguimento_path = os.path.join(folder_path, seguimento_file)

            self.update_status(f"Reading '{loading_file}'...")
            df_loading = pd.read_excel(loading_path, dtype=str)
            df_loading.rename(columns=column_mapping, inplace=True)

            mapped_cols = [col for col in df_loading.columns if col in column_mapping.values()]
            df_loading = df_loading[mapped_cols]

            if df_loading.empty:
                messagebox.showwarning("Warning", "No valid mapped data found in the loading file. Nothing to append.")
                self.update_status("Warning: No valid data found in loading file.")
                return

            for col in date_columns:
                if col in df_loading.columns:
                    df_loading[col] = df_loading[col].apply(self.parse_date)

            self.progress["value"] = 40
            self.update_status("Data loaded and parsed. Preparing to update 'seguimiento' file...")
            self.root.update_idletasks()

            backup_path = seguimento_path.replace(".xlsx", "_backup.xlsx")
            shutil.copy(seguimento_path, backup_path)
            self.update_status(f"Backup created: {os.path.basename(backup_path)}")

            wb = load_workbook(seguimento_path)
            if "Seguimento" not in wb.sheetnames:
                messagebox.showerror("Error", "'Seguimento' sheet not found in the 'seguimiento' Excel file. Please ensure the sheet name is correct.")
                self.update_status("Error: 'Seguimento' sheet not found.")
                return

            ws = wb["Seguimento"]

            headers_excel = [cell.value for cell in ws[1]]
            col_mapping_excel = {
                col: headers_excel.index(col) + 1
                for col in df_loading.columns if col in headers_excel
            }

            if not col_mapping_excel:
                messagebox.showerror("Error", "None of the required columns for mapping were found in the 'Seguimento' sheet's header row.")
                self.update_status("Error: No mapped columns found in sheet.")
                return

            # --- Start Table Handling ---
            existing_table_details = None
            if ws.tables:
                table_name = list(ws.tables.keys())[0]
                existing_table = ws.tables[table_name]
                
                existing_table_details = {
                    'name': table_name,
                    'ref': existing_table.ref,
                    'style_name': "TableStyleMedium9",
                    'show_header_row': True,
                    'show_first_column': False,
                    'show_last_column': False,
                    'show_row_stripes': True,
                    'show_column_stripes': False
                }

                if existing_table.tableStyleInfo:
                    style_info = existing_table.tableStyleInfo
                    try:
                        existing_table_details['style_name'] = getattr(style_info, 'name', "TableStyleMedium9")
                        existing_table_details['show_header_row'] = getattr(style_info, 'showHeaderRow', True)
                        existing_table_details['show_first_column'] = getattr(style_info, 'showFirstColumn', False)
                        existing_table_details['show_last_column'] = getattr(style_info, 'showLastColumn', False)
                        existing_table_details['show_row_stripes'] = getattr(style_info, 'showRowStripes', True)
                        existing_table_details['show_column_stripes'] = getattr(style_info, 'showColumnStripes', False)
                    except Exception as style_error:
                        self.update_status(f"Warning: Could not fully retrieve table style info. Using defaults.")
                        messagebox.showwarning("Table Style Warning", 
                                               f"Could not fully retrieve table style info. Using default style. Error: {style_error}")
                
                del ws.tables[table_name]
                ws._tables.pop(existing_table.id, None) 

            # Store existing formulas from the last data row (still needed as a Python fallback)
            formula_templates = {}
            last_formula_row = 0
            if existing_table_details:
                original_table_end_cell = existing_table_details['ref'].split(':')[1]
                last_formula_row = int(''.join(filter(str.isdigit, original_table_end_cell)))
            else:
                for row in range(ws.max_row, 0, -1):
                    if ws.cell(row=row, column=1).value not in [None, ""]:
                        last_formula_row = row
                        break
            
            if last_formula_row > 0:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=last_formula_row, column=col)
                    if cell.data_type == 'f':
                        formula_templates[col] = cell.value

            current_sheet_last_data_row = ws.max_row
            for row in range(ws.max_row, 0, -1):
                if ws.cell(row=row, column=1).value not in [None, ""]:
                    current_sheet_last_data_row = row
                    break

            start_row_for_new_data = current_sheet_last_data_row + 1

            self.progress["value"] = 70
            self.update_status(f"Appending {len(df_loading)} rows...")
            self.root.update_idletasks()

            # Write data
            for i, row in df_loading.iterrows():
                target_row_num = start_row_for_new_data + i
                for col_name, col_index in col_mapping_excel.items():
                    value = row[col_name]
                    cell = ws.cell(row=target_row_num, column=col_index, value=value)
                    
                    if col_name in date_columns and isinstance(value, datetime):
                        cell.number_format = numbers.FORMAT_DATE_DMYSLASH
                    elif col_name in date_columns and value is None:
                        cell.value = ""

                if last_formula_row > 0:
                    for col, formula in formula_templates.items():
                        adjusted_formula = formula.replace(str(last_formula_row), str(target_row_num))
                        ws.cell(row=target_row_num, column=col, value=adjusted_formula)

            new_sheet_max_row = ws.max_row

            # --- Re-create/Update Table (Python part) ---
            if existing_table_details:
                original_ref_parts = existing_table_details['ref'].split(':')
                start_cell_ref_str = original_ref_parts[0]
                # We need to preserve the original header row, not just the start_row_for_new_data
                original_table_header_row = int(''.join(filter(str.isdigit, start_cell_ref_str))) 
                end_col_letter = ''.join(filter(str.isalpha, original_ref_parts[1]))

                new_table_ref = f"{start_cell_ref_str}:{end_col_letter}{new_sheet_max_row}"

                new_table = Table(displayName=existing_table_details['name'], ref=new_table_ref)
                
                style = TableStyleInfo(name=existing_table_details['style_name'])
                style.showHeaderRow = existing_table_details['show_header_row']
                style.showFirstColumn = existing_table_details['show_first_column']
                style.showLastColumn = existing_table_details['show_last_column']
                style.showRowStripes = existing_table_details['show_row_stripes']
                style.showColumnStripes = existing_table_details['show_column_stripes']
                
                new_table.tableStyleInfo = style

                ws.add_table(new_table)
            else:
                messagebox.showwarning("Warning", "No Excel Table found in 'Seguimento' sheet. Data appended as raw rows. Consider adding a table manually for better formatting.")
                self.update_status("Warning: No Excel table found.")

            self.progress["value"] = 90
            self.update_status("Saving updated Excel file...")
            self.root.update_idletasks()

            # Save the workbook (before triggering VBA)
            wb.save(seguimento_path)
            wb.close()

            # --- Trigger VBA Macro in Run_Update.xlsm ---
            if VBA_AUTOMATION_AVAILABLE:
                excel = None # Initialize excel object
                try:
                    excel = win32.Dispatch('Excel.Application')
                    excel.Visible = False # Keep Excel hidden
                    excel.DisplayAlerts = False # Prevent pop-ups from Excel

                    run_update_workbook_path = os.path.join(Update_file_path, "Run_Update.xlsm")
                    
                    if not os.path.exists(run_update_workbook_path):
                        messagebox.showerror("VBA Error", f"Run_Update.xlsm not found at: {run_update_workbook_path}\nEnsure it's in a 'Dados' subfolder relative to the executable.")
                        self.update_status("Error: Run_Update.xlsm not found.")
                        return 
                        
                    # Check if Run_Update.xlsm is already open
                    macro_workbook = None
                    for wb_open in excel.Workbooks:
                        if os.path.samefile(wb_open.FullName, run_update_workbook_path):
                            macro_workbook = wb_open
                            break
                    
                    if macro_workbook is None: # If not already open, open it
                        macro_workbook = excel.Workbooks.Open(run_update_workbook_path)
                    
                    self.update_status("Running VBA macro for formatting and validation...")
                    # Run the macro (e.g., named "ApplyFormattingToSeguimentoFile")
                    # The macro will find and open the 'seguimento' file itself.
                    excel.Application.Run("ApplyFormattingToSeguimentoFile") 
                    
                    # Close the macro workbook only if we opened it.
                    # This is important to prevent accidental closing if it was already open (e.g., if it's a Personal.xlsb)
                    if not macro_workbook.Name == "PERSONAL.XLSB": # Personal.xlsb is a special case
                        macro_workbook.Close(SaveChanges=False) 
                    
                    excel.DisplayAlerts = True # Re-enable alerts
                    excel.Quit() # Quit Excel application
                    
                    messagebox.showinfo("Success", 
                                        f"Successfully appended {len(df_loading)} rows.\n"
                                        "Formatting and Validation refreshed by VBA macro.\n"
                                        f"Backup saved as:\n{os.path.basename(backup_path)}")
                    self.update_status("Processing complete! VBA macro executed.")

                except Exception as vba_error:
                    messagebox.showwarning("VBA Automation Error", 
                                           f"Python could not trigger VBA macro. This might affect advanced formatting and data validation. Error: {vba_error}\n"
                                           "Please check if 'pywin32' is installed and Excel is not running an interfering process.")
                    # If VBA fails, provide general success message for Python part
                    messagebox.showinfo("Success (without VBA)", 
                                         f"Successfully appended {len(df_loading)} rows.\n"
                                         "However, Conditional Formatting and Data Validation were NOT refreshed (VBA automation failed).\n"
                                         f"Backup saved as:\n{os.path.basename(backup_path)}")
                    self.update_status("Processing complete with VBA error.")
                finally:
                    # Ensure Excel is properly closed in case of errors
                    if excel is not None:
                        try:
                            excel.Quit()
                        except:
                            pass # Ignore if Excel is already closed
            else:
                messagebox.showinfo("Success", 
                                     f"Successfully appended {len(df_loading)} rows.\n"
                                     "Conditional Formatting and Data Validation were NOT refreshed (VBA automation not available).\n"
                                     f"Backup saved as:\n{os.path.basename(backup_path)}")
                self.update_status("Processing complete (VBA not available).")


        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            self.update_status(f"Error: {e}")
        finally:
            self.progress["value"] = 0 # Reset progress bar
            self.btn_browse.config(state=tk.NORMAL) # Re-enable button


# Launch GUI
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMergerApp(root)
    root.mainloop()